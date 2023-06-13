from openpyxl import workbook,load_workbook
from openpyxl.styles import *
import pandas as pd

#颜色
Color(index=0) # 根据索引进行填充
Color(rgb='00000000') # 根据rgb值进行填充
# index 
COLOR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333', '404040' ,"ED1C24", #60-63
    'A6A6A6'
)
BLACK = COLOR_INDEX[0]
WHITE = COLOR_INDEX[1]
RED = COLOR_INDEX[65]
DARKRED = COLOR_INDEX[8]
BLUE = COLOR_INDEX[4]
DARKBLUE = COLOR_INDEX[12]
GREEN = COLOR_INDEX[3]
DARKGREEN = COLOR_INDEX[9]
YELLOW = COLOR_INDEX[5]
DARKYELLOW = COLOR_INDEX[19]
GRAY = COLOR_INDEX[64]
BORDER=COLOR_INDEX[66]

# 填充色
PatternFill(patternType='solid',fgColor=Color(), bgColor=Color())
# fgColor   前景色
# bgColor   后景色
# 参数可选项
patternType = {'darkDown', 'darkUp', 'lightDown', 'darkGrid', 'lightVertical', 
               'solid', 'gray0625', 'darkHorizontal', 'lightGrid', 'lightTrellis', 
               'mediumGray', 'gray125', 'darkGray', 'lightGray', 'lightUp', 
               'lightHorizontal', 'darkTrellis', 'darkVertical'}

def generate_table(tabel_name,wb):
    wb.create_sheet(tabel_name)
    sheet_test = wb[tabel_name]
    sheet_test.page_margins.left=0.25
    sheet_test.page_margins.right=0.25
    sheet_test.page_margins.top = 0.3
    sheet_test.page_margins.bottom = 0.5
    # title
    sheet_test.merge_cells("A1:D1")
    sheet_test["A1"].value = "Provision of Personal Data to Third Parties"
    # col & row 
    sheet_test.row_dimensions[1].height = 22
    sheet_test.column_dimensions["A"].width=18    #21
    sheet_test.column_dimensions["B"].width=24       #28
    sheet_test.column_dimensions["C"].width=22           #24
    sheet_test.column_dimensions["D"].width=27             #25
    # font
    sheet_test["A1"].font = Font(name="Calibri",size=13,color=WHITE,b=True,i=False)
    sheet_test["A1"].fill = PatternFill(patternType='solid',fgColor=Color(RED), bgColor=Color())
    sheet_test["A1"].alignment = Alignment(horizontal="left",vertical="center")
    # wb.save("test.xlsx")  


def box_info(star_line,wb,table_name,info):

    sheet_test = wb[table_name]
    box_line_row1 = star_line+1
    box_line_row2 = star_line+2
    box_line_row3 = star_line+3
    box_line_row4 = star_line+4
    box_line_row5 = star_line+5
    box_line_row6 = star_line+6
    box_line_row7 = star_line+7
    box_line_row8 = star_line+8

    sheet_test["A"+str(box_line_row2)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row1)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["C"+str(box_line_row2)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row3)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["C"+str(box_line_row3)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row4)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["C"+str(box_line_row4)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row5)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["C"+str(box_line_row5)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row6)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["C"+str(box_line_row6)].fill = PatternFill(patternType='solid',fgColor=Color(WHITE), bgColor=Color())
    sheet_test["A"+str(box_line_row4)].border = Border(top=Side(style="thin",color=BORDER))
    sheet_test["C"+str(box_line_row4)].border = Border(top=Side(style="thin",color=BORDER))
    sheet_test["B"+str(box_line_row6)].border = Border(left=Side(style="thin",color=BORDER),right=Side(style="thin",color=BORDER),top=Side(style="thin",color=BORDER),bottom=Side(style="thin",color=BORDER))

    #title Name&ID

    sheet_test.merge_cells("A"+str(box_line_row1)+":D"+str(box_line_row1))
    row_2 = sheet_test.row_dimensions[box_line_row1]
    row_2.height = 20
    sheet_test["A"+str(box_line_row1)].value = info[0]
    sheet_test["A"+str(box_line_row1)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["A"+str(box_line_row1)].font = Font(name="Calibri",size=13,color=BLACK,b=True,i=False)

    sheet_test.merge_cells("A"+str(box_line_row2)+":B"+str(box_line_row2))
    sheet_test.merge_cells("C"+str(box_line_row2)+":D"+str(box_line_row2))
    row_3 = sheet_test.row_dimensions[box_line_row2]
    row_3.height = 15
    sheet_test["A"+str(box_line_row2)].value = "Company URL or Privacy Policy"
    sheet_test["C"+str(box_line_row2)].value = "Data Location"
    sheet_test["A"+str(box_line_row2)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["C"+str(box_line_row2)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["A"+str(box_line_row2)].font = Font(name="Calibri",size=10,color=BLACK,b=True,i=False)
    sheet_test["C"+str(box_line_row2)].font = Font(name="Calibri",size=10,color=BLACK,b=True,i=False)

    sheet_test.merge_cells("A"+str(box_line_row3)+":B"+str(box_line_row3))
    sheet_test.merge_cells("C"+str(box_line_row3)+":D"+str(box_line_row3))
    row_4 = sheet_test.row_dimensions[box_line_row3]
    row_4.height = 35
    sheet_test["A"+str(box_line_row3)].value = info[1]
    sheet_test["C"+str(box_line_row3)].value = info[2]
    sheet_test["A"+str(box_line_row3)].alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)
    sheet_test["C"+str(box_line_row3)].alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)
    sheet_test["A"+str(box_line_row3)].font = Font(name="Calibri",size=10,color=GRAY,b=False,i=False)
    sheet_test["C"+str(box_line_row3)].font = Font(name="Calibri",size=10,color=GRAY,b=False,i=False)

    sheet_test.merge_cells("A"+str(box_line_row4)+":B"+str(box_line_row4))
    sheet_test.merge_cells("C"+str(box_line_row4)+":D"+str(box_line_row4))
    row_5 = sheet_test.row_dimensions[box_line_row4]
    row_5.height = 15
    sheet_test["A"+str(box_line_row4)].value = "Purpose of Data Share"
    sheet_test["C"+str(box_line_row4)].value = "Downside of opting out"
    sheet_test["A"+str(box_line_row4)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["C"+str(box_line_row4)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["A"+str(box_line_row4)].font = Font(name="Calibri",size=10,color=BLACK,b=True,i=False)
    sheet_test["C"+str(box_line_row4)].font = Font(name="Calibri",size=10,color=BLACK,b=True,i=False)

    sheet_test.merge_cells("A"+str(box_line_row5)+":B"+str(box_line_row5))
    sheet_test.merge_cells("C"+str(box_line_row5)+":D"+str(box_line_row5))
    row_6 = sheet_test.row_dimensions[box_line_row5]
    row_6.height = 37
    sheet_test["A"+str(box_line_row5)].value = info[3]
    sheet_test["C"+str(box_line_row5)].value = info[4]
    sheet_test["A"+str(box_line_row5)].alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)
    sheet_test["C"+str(box_line_row5)].alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)
    sheet_test["A"+str(box_line_row5)].font = Font(name="Calibri",size=10,color=GRAY,b=False,i=False)
    sheet_test["C"+str(box_line_row5)].font = Font(name="Calibri",size=10,color=GRAY,b=False,i=False)

    sheet_test.merge_cells("B"+str(box_line_row6)+":D"+str(box_line_row6))
    row_7 = sheet_test.row_dimensions[box_line_row6]
    row_7.height = 61
    sheet_test["A"+str(box_line_row6)].value = "Data Items Shared"
    sheet_test["A"+str(box_line_row6)].font = Font(name="Calibri",size=10,color=BLACK,b=True,i=False)
    sheet_test["B"+str(box_line_row6)].value = info[5]
    sheet_test["B"+str(box_line_row6)].font = Font(name="Calibri",size=10,color=GRAY,b=False,i=False)
    sheet_test["A"+str(box_line_row6)].alignment = Alignment(horizontal="left",vertical="top")
    sheet_test["B"+str(box_line_row6)].alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)

    sheet_test.merge_cells("A"+str(box_line_row7)+":D"+str(box_line_row7))
    row_8 = sheet_test.row_dimensions[box_line_row7]
    row_8.height = 5

    sheet_test.merge_cells("A"+str(box_line_row8)+":D"+str(box_line_row8))
    row_9 = sheet_test.row_dimensions[box_line_row8]
    row_9.height = 7
    sheet_test["A"+str(box_line_row8)].fill = PatternFill(patternType="solid",fgColor=Color(RED))
    

def generate_data(Name,DF):
    df = DF[DF[Name] == "Y"]
    df_sorted = df.sort_values(by=["ID#"]).fillna(value=0)
    data_list = []
    for i in df_sorted.iterrows():

        # ID generate
        ID = str(i[1][0]) +" "+str(i[1][1])

        #Link or URL
        if i[1]["Link"] == 0:
            if i[1]["Privacy Policy"] ==0:
                Link = ""
            else:
                Link = i[1]["Privacy Policy"]
        elif i[1]["Link"] == "No official website":
            if i[1]["Privacy Policy"] ==0:
                Link = ""
            else:
                Link = i[1]["Privacy Policy"]
        else:
            Link = i[1]["Link"]

        #Data location
        if i[1]["Data location"] ==0 :
            Data_location = ""
        else:
            Data_location = i[1]['Data location']

        # Purpose of data share
        if i[1]["Purpose"] == 0:
            Purpose = ''
        else:
            Purpose = i[1]['Purpose']

        #Downside of opting out
        if i[1]["Possible downsides of \'opting out\'"] == 0:
            Dooo = ""
        else:
            Dooo = i[1]["Possible downsides of \'opting out\'"]

        # Data Items shared
        item_list = ["User SIS IDs","Student Name","Staff Name","Parent Name","Student Email","Staff Email","Parent Email","Student Contact (other)",
                     "Staff Contact (other)","Parent Contact (other)","Parent Relationship","Student ID (Personal Identification)","Staff ID (Personal Identification)",
                    "Parent ID (Personal Identification)","Student SIS Memberships","Staff SIS Memberships","Student DOB","Staff DOB","Parent DOB","Student Assessment Data"
                ,"Student Address","Student Phone","Parent Phone","Student Image","Staff Image","Parent Image","Role","Gender","Transactional Profile","Transactional Communication",
                    "Transactional Assessment","Transactional Behaviours","Sensitive Safeguarding","Sensitive Learner Profile","Sensitive Medical","Student Ethnicity",
                    "Student Languages","Other Academic"]
        Items=''
        for k in item_list:
            if i[1][k] == "Y":
                Items += k
                Items += ","
        final_info = [ID,Link,Data_location,Purpose,Dooo,Items]
        data_list.append(final_info)
    return data_list


#主逻辑
DF_original = pd.read_excel("DAR 0605.xlsx")
#整理总表
title = DF_original.iloc[0]
DF_original.columns = list(title)
DF = DF_original.drop(labels=0,axis=0)
# DF_list = ["DCB"]
DF_list = ['DCB','DCSG','DCSL','DCSPD','DCSPX','DCSZ','DEGT','DEMH','DEXA','DHSZ','DHZH','DUSZ','HQ']
wb = load_workbook("DAR 0605.xlsx")
for i in DF_list:
    #生成一张张学校的表格
    generate_table(i,wb)
    #生成数据
    data = generate_data(i,DF)
    #生成表格
    for index , info in enumerate(data):
        star_line = index * 8 + 1
        box_info(star_line,wb,i,info)
    print(i,"'s report generate successfully")
# wb.remove_sheet(["Sheet1"])
wb.save("Final_Report.xlsx")  
print("All reports generated successfully")

   






